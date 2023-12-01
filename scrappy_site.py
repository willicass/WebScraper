from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from credenciais import usuario, senha
from time import sleep
import openpyxl
import smtplib
from email.message import EmailMessage
import re

nome_item = 'Fogo Rápido'
lista_loja_item = []
lista_preco_item = []
lista_quantidade = []


def email_usuario():
    while True:
        email = input('Digite seu email: ')
        email = email.lower()
        padrao = re.search(r'[a-zA-Z0-9_-]+@[a-zA-Z0-9]+\.[a-zA-Z]{1,3}$', email)
        if padrao:
            print('\u001b[32mEmail válido\u001b[0m')
            return email
        else:
            print('Digite um email válido!!!')
            break 

def pesquisar_item():
    driver = webdriver.Chrome()
    driver.set_window_size(800, 700)
    link = 'https://ragnatales.com.br/market'
    print(driver.title)
    lista_loja_item= []
    lista_preco_item = []
    lista_quantidade  = []
    driver.get(link)
    barra_pesquisa = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, f'//*[@id="app"]/div/div/div[2]/div/div[2]/div[1]/label/div/div[1]/input')))
    barra_pesquisa.send_keys(nome_item)
    botao_pesquisa = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, f'//*[@id="app"]/div/div/div[2]/div/div[2]/div[1]/label/div/div[2]/button/div')))
    
    sleep(2)
    botao_pesquisa.click()
    sleep(8)

    return driver

def coletando_dados(driver):
    valido = False
    validacao = []
    total_paginas = 0
    print('Coletando dados ...')
    for item in range(1, 21):
        try:
            localizacao_item = WebDriverWait(driver, 12).until(EC.presence_of_element_located((By.XPATH,
                f'//*[@id="app"]/div/div/div[2]/div/div[2]/div[2]/div/div/div/div/div/table/tbody/tr[{item}]/td[4]/div/div[1]/button/div/div/div[2]/div[2]'))).text
            preco_item = WebDriverWait(driver, 12).until(EC.presence_of_element_located((By.XPATH,
                f'//*[@id="app"]/div/div/div[2]/div/div[2]/div[2]/div/div/div/div/div/table/tbody/tr[{item}]/td[3]/div/span'))).text
            quantidade = WebDriverWait(driver, 12).until(EC.presence_of_element_located((By.XPATH,
                f'//*[@id="app"]/div/div/div[2]/div/div[2]/div[2]/div/div/div/div/div/table/tbody/tr[{item}]/td[2]/div/span'))).text
            chave = quantidade + localizacao_item + preco_item
            if chave not in validacao:
                validacao.append(chave)
                lista_loja_item.append(localizacao_item)
                lista_preco_item.append(preco_item)
                lista_quantidade.append(quantidade)
                print(chave)
        except:
            return True
    
    return lista_loja_item, lista_preco_item, lista_quantidade


def passar_pagina(driver):
    for paginas in range(4, 12):
        url = driver.current_url
        coletando_dados(driver) 
        proxima_pagina = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
        f'//*[@id="app"]/div/div/div[2]/div/div[2]/div[2]/div/div/div/div/div/table/tfoot/tr/td/div/button[{paginas}]')))
        proxima_pagina.click()
        print('\u001b[32mPassando de página\u001b[0m')
        sleep(4)
        url_atual = driver.current_url
        if url == url_atual:
            return True


def criar_planilha(lista_preco_item = lista_preco_item, lista_loja_item=lista_loja_item , lista_quantidade=lista_quantidade, nome_item = nome_item):
    index = 2
    planilha = openpyxl.Workbook()
    itens = planilha['Sheet']
    itens.title = 'Preço e localização de itens'
    itens['A1'] = 'Nome'
    itens['B1'] = 'Preço'
    itens['C1'] = 'Localização'
    itens['D1'] = 'Quantidade'
    for preco, localizacao, quantidade in zip(lista_preco_item, lista_loja_item, lista_quantidade):
        itens.cell(column=2, row=index, value=preco)
        itens.cell(column=3, row=index, value=localizacao)
        itens.cell(column=4, row=index, value=quantidade)
        index += 1
    for index in range(len(lista_loja_item) + 2):
        if index == 0:
            index = 2
        itens.cell(column=1, row=index, value=nome_item)
    planilha.save("planilha_Ragnarok.xlsx")
    print(f'\u001b[32m{"Planilha criada com sucesso"}\u001b[0m')

def enviar_email_cliente(email):
    endereco = usuario
    msg = EmailMessage()
    msg['Subject'] = 'planilha Ragnarok'
    msg['From'] = usuario
    msg['To'] = email
    msg.set_content('Aqui ta seus itens meu chapa')
    arquivos = ["planilha_Ragnarok.xlsx"]
    for arquivo in arquivos:
        with open(arquivo, 'rb') as arq:
            dados = arq.read()
            nome_arquivo = arq.name
        msg.add_attachment(dados, maintype='application',
                            subtype='octet-stream', filename=nome_arquivo)
    server = smtplib.SMTP('imap.gmail.com', 587)
    server.ehlo()
    server.starttls()
    server.login(endereco, senha, initial_response_ok=True)
    server.send_message(msg)
    print(f'\u001b[32m{"Enviando email para para destinatario"}\u001b[0m')
    server.quit()

def iniciar():
    email = email_usuario()
    driver = pesquisar_item()
    coletando_dados(driver)
    passar_pagina(driver)
    criar_planilha()
    enviar_email_cliente(email)

iniciar()